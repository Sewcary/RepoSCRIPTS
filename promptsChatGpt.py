import openpyxl
import os
import openai

# Récupérer la clé API OpenAI depuis une variable d'environnement
openai.api_key = "sk-...LeHf"


# Ouvrir le fichier Excel
wb = openpyxl.load_workbook('fichier_excel.xlsx')

# Sélectionner la première feuille
sheet = wb.active

# Parcourir toutes les cellules de la première colonne
for row in sheet.iter_rows():
    cell = row[0]
    prompt = cell.value.strip()
    if prompt:
        # Appeler l'API ChatGPT
        response = openai.Completion.create(
            engine="davinci",
            prompt=prompt,
            max_tokens=1024,
            n=1,
            stop=None,
            temperature=0.5,
        )

        # Récupérer la réponse et l'enregistrer dans la cellule de la deuxième colonne
        answer = response.choices[0].text.strip()
        sheet.cell(row=cell.row, column=2).value = answer

# Enregistrer le fichier modifié
wb.save('fichier_excel_modifié.xlsx')

response = requests.get('https://api.openai.com/v1/engines/davinci/completions', verify=False)